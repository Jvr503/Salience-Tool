# Salience Tool — Propellic
# Created by Javier Hernandez

import base64
import json
import io
import os
import re
import requests
import pandas as pd
import numpy as np
import streamlit as st
import streamlit.components.v1 as components
from datetime import datetime
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from google.cloud import language_v1
import anthropic
import db
import auth

load_dotenv()

# ── Google Cloud credentials (write JSON from secret if running on Streamlit Cloud) ──
def _setup_google_credentials():
    creds_path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "")
    if creds_path and os.path.exists(creds_path):
        return  # local dev, file already exists
    creds_json = os.getenv("GOOGLE_CREDENTIALS_JSON") or st.secrets.get("GOOGLE_CREDENTIALS_JSON", "")
    if creds_json:
        import tempfile
        tmp = tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False)
        tmp.write(creds_json)
        tmp.flush()
        os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = tmp.name

_setup_google_credentials()

# ── DB init + admin seed ───────────────────────────────────────────────────────
db.init_db()
db.seed_admin("javier.hernandez@propellic.com", "Javier Hernandez")

CLAUDE_MODEL = os.getenv("CLAUDE_MODEL", "claude-opus-4-6")

def _get_anthropic_key() -> str:
    """DB key takes precedence; falls back to env var then st.secrets."""
    from_db = db.get_api_key("anthropic")
    if from_db:
        return from_db
    return os.getenv("ANTHROPIC_API_KEY") or st.secrets.get("ANTHROPIC_API_KEY", "")

ANTHROPIC_API_KEY = _get_anthropic_key()

# ── Helpers ───────────────────────────────────────────────────────────────────

def first_sentence(text: str) -> str:
    text = re.sub(r"\s+", " ", (text or "").strip())
    if not text:
        return ""
    parts = re.split(r"(?<=[.!?])\s+", text, maxsplit=1)
    return parts[0].strip()


def fetch_page_elements(url: str) -> dict:
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X) AppleWebKit/537.36 "
                      "(KHTML, like Gecko) Chrome Safari"
    }
    r = requests.get(url, headers=headers, timeout=15)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")

    meta_title = soup.title.get_text(strip=True) if soup.title else ""

    meta_desc = ""
    md = soup.find("meta", attrs={"name": re.compile(r"^description$", re.I)})
    if md and md.get("content"):
        meta_desc = md.get("content", "").strip()

    h1_text = ""
    h1 = soup.find("h1")
    if h1:
        h1_text = h1.get_text(" ", strip=True)

    after_h1_text = ""
    if h1:
        p = h1.find_next("p")
        if p:
            after_h1_text = p.get_text(" ", strip=True)

    fs = first_sentence(after_h1_text)

    return {
        "meta_title": meta_title,
        "meta_description": meta_desc,
        "h1": h1_text,
        "first_sentence_after_h1": fs,
        "prefill_original": "\n\n".join([t for t in [h1_text, fs] if t]),
    }


def analyze_text_salience(text: str) -> dict:
    client = language_v1.LanguageServiceClient()
    document = language_v1.Document(content=text, type_=language_v1.Document.Type.PLAIN_TEXT)
    response = client.analyze_entities(document=document)
    result = {}
    for entity in response.entities:
        result[entity.name] = {
            "Type": language_v1.Entity.Type(entity.type_).name,
            "Salience": float(entity.salience),
        }
    return result


def top_entity(entities: dict):
    if not entities:
        return None, None
    top = max(entities.items(), key=lambda x: x[1]["Salience"])
    return top[0], top[1]["Salience"]


def wc(text: str) -> int:
    return len(text.split()) if text.strip() else 0


def score_color(val: float) -> str:
    if val >= 0.60: return "#16a34a"
    if val >= 0.40: return "#ca8a04"
    if val >= 0.20: return "#ea580c"
    return "#dc2626"


def style_score(val):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return ""
    if val >= 0.60: return "color: #16a34a; font-weight: 600"
    if val >= 0.40: return "color: #ca8a04; font-weight: 600"
    if val >= 0.20: return "color: #ea580c; font-weight: 600"
    return "color: #dc2626; font-weight: 600"


def copy_button(text: str, key: str = "") -> None:
    """Renders a clipboard copy button via JS component."""
    if not text.strip():
        return
    # base64-encode the text so it's safe to embed directly in an HTML attribute
    # (avoids quote-escaping issues that break onclick when text contains " or ')
    b64 = base64.b64encode(text.encode("utf-8")).decode("ascii")
    components.html(
        f"""<button
          onclick="var t=atob('{b64}');navigator.clipboard.writeText(t).then(()=>{{
            this.textContent='&#10003; Copied!';
            setTimeout(()=>this.textContent='&#128203; Copy',2000);
          }})"
          style="background:#E21A6B;color:white;border:none;border-radius:6px;
                 padding:5px 14px;cursor:pointer;font-size:12px;font-weight:bold;
                 font-family:Montserrat,sans-serif;">
          &#128203; Copy
        </button>""",
        height=38,
    )


# ── Auth gate ─────────────────────────────────────────────────────────────────
_current_user = auth.require_auth()

# Refresh API key each page load (picks up changes saved in admin panel)
ANTHROPIC_API_KEY = _get_anthropic_key()

# ── Credentials ───────────────────────────────────────────────────────────────

cred = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS")
if not cred:
    st.error("GOOGLE_APPLICATION_CREDENTIALS is not set. Add it to ~/.zprofile or .env, then restart.")
    st.stop()

cred_path = os.path.expanduser(cred)
if not os.path.exists(cred_path):
    st.error(f"GOOGLE_APPLICATION_CREDENTIALS points to a missing file: {cred}")
    st.stop()

# ── Claude helpers ─────────────────────────────────────────────────────────────

@st.cache_resource
def get_claude_client(api_key: str) -> anthropic.Anthropic:
    return anthropic.Anthropic(api_key=api_key)


_SHORT_ELEMENTS = {"H1", "Meta title"}


def generate_optimized_text(original: str, target_entity: str, source_label: str = "Pasted text") -> str:
    client = get_claude_client(ANTHROPIC_API_KEY)

    if source_label in _SHORT_ELEMENTS:
        word_cnt = len(original.split())
        prompt = f"""You are optimizing a {source_label} for Google Cloud Natural Language API entity salience.

Goal: Rewrite this {source_label} so that "{target_entity}" is the #1 entity by salience.

Rules:
1) This is a {source_label} — keep it as a short heading. Do NOT write a sentence or paragraph.
2) Keep the output at roughly {word_cnt} words (±2 words maximum). Never expand it.
3) Place "{target_entity}" as the first or second word of the heading.
4) Remove or demote any competing named entities — replace them with generic descriptors if needed.
5) Return ONLY the rewritten {source_label}. No punctuation at the end unless the original had it. No explanation.

Original {source_label}:
{original}""".strip()
    else:
        prompt = f"""You are optimizing text for Google Cloud Natural Language API entity salience.

Goal: The entity "{target_entity}" must be the #1 entity by salience when analyzed with Google NLP.

Rewrite rules:
1) Put "{target_entity}" in the first sentence, within the first 8–10 words.
2) Make "{target_entity}" the grammatical subject of most sentences (active voice).
3) Use the exact string "{target_entity}" multiple times naturally.
4) Demote competing entities — rewrite them as generic references after the first mention.
5) Do not introduce new named entities not in the original text.
6) Same length or shorter — do NOT expand the original text.

Return ONLY the rewritten text. No explanation. No **.

Original text:
{original}""".strip()

    message = client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=512,
        messages=[{"role": "user", "content": prompt}],
    )
    return message.content[0].text.strip()


def suggest_entity_for_keyword(entities: list, keyword: str) -> str:
    client = get_claude_client(ANTHROPIC_API_KEY)
    entity_list = ", ".join(f'"{e}"' for e in entities[:20])
    prompt = (
        f'You are an SEO specialist. A page contains these named entities: {entity_list}\n\n'
        f'A user wants to optimize this page to rank for the keyword: "{keyword}"\n\n'
        f'Which SINGLE entity from the list should be made the most salient (dominant subject) '
        f'to best align this page with that keyword from an SEO perspective? '
        f'Choose the entity that most directly represents the searcher\'s intent behind "{keyword}" '
        f'— not necessarily the brand or the most prominent entity currently on the page.\n\n'
        f'Reply with ONLY the entity name, exactly as written in the list above. No explanation.'
    )
    message = client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=50,
        messages=[{"role": "user", "content": prompt}],
    )
    return message.content[0].text.strip().strip('"')


# ── Branding / CSS ─────────────────────────────────────────────────────────────

st.markdown(
    """
    <style>
    html, body, [class*="css"] { font-family: 'Montserrat', sans-serif; }
    .stButton>button {
        background-color: #E21A6B; color: white; font-weight: bold;
        border: none; border-radius: 8px; padding: 0.6em 1.4em;
    }
    .stButton>button:hover { background-color: #c0175d; }
    </style>
    """,
    unsafe_allow_html=True,
)

st.image("propellic-logo-png.png", width=180)
st.markdown(
    """
    <div style="text-transform:uppercase;color:#E21A6B;font-weight:bold;font-size:14px;margin-bottom:0.5rem;">
      Salience Analyzer
    </div>
    <h1 style='color:white;margin-top:0;'>Text Analysis with Google NLP</h1>
    """,
    unsafe_allow_html=True,
)

# ── Session state ──────────────────────────────────────────────────────────────

defaults: dict = {
    # URL tab
    "page_url": "",
    "url_meta_title": "",
    "url_meta_description": "",
    "url_h1": "",
    "url_first_sentence": "",
    "url_element_scores": {},   # {key: {"entity": str, "score": float}}
    # Paste tab
    "original_text": "",
    # Variations
    "variation_text_1": "",
    "variation_text_2": "",
    "target_entity_1": "",
    "target_entity_2": "",
    "variation_source_1": "Pasted text",
    "variation_source_2": "Pasted text",
    "assign_to": "Variation 1",
    "entity_picker_1": "",
    "entity_picker_2": "",
    # Results
    "display_df": None,
    "claude_error": "",
    "analyze_v1_label": "",
    "analyze_v2_label": "",
    "selected_original_key": "",
    # Keyword suggestion
    "target_keyword": "",
    "keyword_suggestion": "",
    # History
    "session_history": [],
    # Bulk
    "bulk_urls_input": "",
    "bulk_target_entity": "",
    "bulk_results_df": None,
}

for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

if st.session_state.get("assign_to") not in ["Variation 1", "Variation 2"]:
    st.session_state["assign_to"] = "Variation 1"

# ── Source options ─────────────────────────────────────────────────────────────

_SOURCE_OPTIONS = ["Pasted text", "H1", "First sentence after H1", "Meta title", "Meta description"]
_SOURCE_KEY_MAP = {
    "Pasted text": "original_text",
    "H1": "url_h1",
    "First sentence after H1": "url_first_sentence",
    "Meta title": "url_meta_title",
    "Meta description": "url_meta_description",
}

# ── Callbacks ──────────────────────────────────────────────────────────────────

def clear_all():
    skip = {"session_history"}
    for k, v in defaults.items():
        if k not in skip:
            st.session_state[k] = v if not isinstance(v, (dict, list)) else type(v)()


def use_as_original(key: str):
    st.session_state["original_text"] = st.session_state.get(key, "")
    st.session_state["selected_original_key"] = key


def assign_selected_entity():
    df = st.session_state.get("display_df")
    table_state = st.session_state.get("entity_table")
    if df is None or table_state is None:
        return
    rows = table_state.selection.rows
    if not rows:
        return
    selected_entity = df.iloc[rows[0]]["Entity"]
    if st.session_state.get("assign_to") == "Variation 1":
        st.session_state["target_entity_1"] = selected_entity
    else:
        st.session_state["target_entity_2"] = selected_entity


def pick_entity(variation_num: int):
    key = f"entity_picker_{variation_num}"
    val = st.session_state.get(key, "")
    if val:
        st.session_state[f"target_entity_{variation_num}"] = val
        st.session_state[key] = ""   # reset picker to placeholder


def on_generate_with_claude():
    if not ANTHROPIC_API_KEY:
        st.session_state["claude_error"] = "ANTHROPIC_API_KEY is not set. Add it to .env and restart."
        return

    is_v1 = st.session_state["assign_to"] == "Variation 1"
    target = st.session_state["target_entity_1"] if is_v1 else st.session_state["target_entity_2"]
    source_label = st.session_state["variation_source_1"] if is_v1 else st.session_state["variation_source_2"]
    source_key = _SOURCE_KEY_MAP[source_label]
    original = st.session_state.get(source_key, "")

    if not original.strip():
        st.session_state["claude_error"] = (
            f'No text found for "{source_label}". '
            + ("Paste text in the Paste tab first." if source_key == "original_text" else "Load a URL first.")
        )
        return
    if not target.strip():
        st.session_state["claude_error"] = "Enter a target entity first."
        return

    st.session_state["claude_error"] = ""
    try:
        new_text = generate_optimized_text(original, target, source_label)
    except Exception as e:
        st.session_state["claude_error"] = f"Claude request failed: {e}"
        return

    if is_v1:
        st.session_state["variation_text_1"] = new_text
    else:
        st.session_state["variation_text_2"] = new_text

    st.toast("Text generated successfully.")


def generate_onpage_recommendation(element_type: str, original: str, primary_kw: str, secondary_kws: str = "") -> str:
    """Generate an optimized rewrite for an on-page element, or return 'leave as is'."""
    if not original.strip():
        return ""
    client = get_claude_client(ANTHROPIC_API_KEY)
    sec_ctx = (f"\nSecondary keywords to incorporate naturally: {secondary_kws}"
               if secondary_kws.strip() else "")
    if element_type in _SHORT_ELEMENTS:
        word_cnt = len(original.split())
        prompt = (
            f'You are an SEO specialist writing on-page recommendations.\n\n'
            f'Element: {element_type}\n'
            f'Primary keyword (target entity): "{primary_kw}"{sec_ctx}\n\n'
            f'Current {element_type}: {original}\n\n'
            f'Rewrite so "{primary_kw}" is the dominant entity.\n\n'
            f'Rules:\n'
            f'1. Keep it a short heading — {word_cnt} words (±2). Never expand to a sentence.\n'
            f'2. Place "{primary_kw}" first or second in the heading.\n'
            f'3. If it already starts with or prominently features "{primary_kw}", reply exactly: leave as is\n'
            f'4. Return ONLY the rewritten heading or "leave as is". No explanation.'
        )
    else:
        char_target = len(original)
        prompt = (
            f'You are an SEO specialist writing on-page recommendations.\n\n'
            f'Element: {element_type}\n'
            f'Primary keyword (target entity): "{primary_kw}"{sec_ctx}\n\n'
            f'Current {element_type}: {original}\n\n'
            f'Rewrite so "{primary_kw}" is the dominant entity for Google NLP salience.\n\n'
            f'Rules:\n'
            f'1. Put "{primary_kw}" in the first 8–10 words.\n'
            f'2. Make "{primary_kw}" the grammatical subject.\n'
            f'3. Keep roughly the same length (target ≈{char_target} chars ± 15).\n'
            f'4. Incorporate secondary keywords naturally if provided.\n'
            f'5. If "{primary_kw}" is already clearly dominant in the first sentence, reply exactly: leave as is\n'
            f'6. Return ONLY the rewritten text or "leave as is". No explanation.'
        )
    msg = client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=300,
        messages=[{"role": "user", "content": prompt}],
    )
    result = msg.content[0].text.strip()
    return "leave as is" if result.lower().startswith("leave as is") else result


def generate_seo_notes(primary_kw: str, title: str, rec_title: str,
                       desc: str, rec_desc: str, h1: str, rec_h1: str,
                       first_sent: str, rec_first: str) -> str:
    """Generate a brief SEO rationale note for the changes made."""
    client = get_claude_client(ANTHROPIC_API_KEY)

    def _changed(orig, rec):
        return rec.strip().lower() != "leave as is" and rec.strip() != orig.strip()

    changed = [name for name, o, r in [
        ("title", title, rec_title), ("meta description", desc, rec_desc),
        ("H1", h1, rec_h1), ("first sentence", first_sent, rec_first),
    ] if _changed(o, r)]

    prompt = (
        f'You are an SEO specialist writing brief notes for an on-page recommendation report.\n\n'
        f'Primary keyword: "{primary_kw}"\n'
        f'Elements changed: {", ".join(changed) if changed else "none (all already well-optimised)"}\n\n'
        f'Existing:\n'
        f'- Page Title: {title[:80]}\n'
        f'- H1: {h1[:60]}\n\n'
        f'Write 1–2 concise sentences: what was changed and the SEO rationale, '
        f'or why elements were left as-is.\n'
        f'Example style: "Align H1 to primary keyword; meta already strong."\n'
        f'Return ONLY the note text.'
    )
    msg = client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=120,
        messages=[{"role": "user", "content": prompt}],
    )
    return msg.content[0].text.strip()


def build_onpage_excel(results: list, client_name: str = "") -> bytes:
    """Build an on-page recommendations Excel file matching the Propellic reference format."""
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "On-Page Recommendations"

    # ── Fills ──
    NAVY         = PatternFill("solid", fgColor="073763")
    COL_ORANGE   = PatternFill("solid", fgColor="E69138")
    PINK         = PatternFill("solid", fgColor="E6235E")
    GREEN        = PatternFill("solid", fgColor="5CBD8D")
    YELLOW       = PatternFill("solid", fgColor="F1C232")
    NOTE_ORANGE  = PatternFill("solid", fgColor="FF9900")

    WHITE_BOLD = Font(bold=True, color="FFFFFF")
    BLUE_LINK  = Font(color="0000FF", underline="single")
    CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT_TOP   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    thin       = Side(style="thin", color="CCCCCC")
    BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Row 1: client name ──
    if client_name.strip():
        cell = ws["A1"]
        cell.value = client_name
        cell.font = Font(bold=True, size=13)
        cell.alignment = CENTER

    # ── Column spec: (header, fill, width) ──
    COLS = [
        ("Approved?",                    NAVY,        13),
        ("Done?",                        NAVY,        11),
        ("Existing URL",                 NAVY,        55),
        ("Primary Keyword",              COL_ORANGE,  30),
        ("KD",                           PINK,         7),
        ("Volume",                       PINK,        12),
        ("Recommended URL",              GREEN,       35),
        ("Secondary Keywords",           PINK,        32),
        ("Page Title",                   PINK,        50),
        ("Length",                       YELLOW,       9),
        ("Recommended Title",            GREEN,       50),
        ("Meta Description",             PINK,        50),
        ("Length",                       YELLOW,       9),
        ("Recommended Meta Description", GREEN,       55),
        ("Length",                       YELLOW,       9),
        ("H1",                           PINK,        35),
        ("Recommended H1",               GREEN,       35),
        ("First Sentence after H1",      PINK,        55),
        ("Length",                       YELLOW,       9),
        ("Recommended First Sentence",   GREEN,       55),
        ("Length",                       YELLOW,       9),
        ("Notes",                        NOTE_ORANGE, 45),
    ]

    # ── Row 2: headers ──
    for ci, (header, fill, width) in enumerate(COLS, 1):
        cell = ws.cell(row=2, column=ci, value=header)
        cell.fill = fill
        cell.font = WHITE_BOLD
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 30

    # ── Data rows ──
    for ri, row in enumerate(results, 3):
        def w(ci, val):
            c2 = ws.cell(row=ri, column=ci, value=val)
            c2.alignment = LEFT_TOP
            c2.border = BORDER
            return c2

        # A,B: leave blank (client fills)
        w(1, ""); w(2, "")

        url_cell = w(3, row.get("Existing URL", ""))
        url_cell.font = BLUE_LINK

        w(4, row.get("Primary Keyword", ""))

        kd = row.get("KD")
        try:    w(5, int(float(kd)) if kd not in (None, "", float("nan")) else "")
        except: w(5, "")

        vol = row.get("Volume")
        try:    w(6, int(float(vol)) if vol not in (None, "", float("nan")) else "")
        except: w(6, "")

        w(7,  row.get("Recommended URL", "") or "Leave as is")
        w(8,  row.get("Secondary Keywords", ""))

        title = row.get("Page Title", "")
        w(9,  title)
        w(10, len(title) if title else "")
        w(11, row.get("Recommended Title", ""))

        desc = row.get("Meta Description", "")
        w(12, desc)
        w(13, len(desc) if desc else "")

        rec_desc = row.get("Recommended Meta Description", "")
        w(14, rec_desc)
        w(15, len(rec_desc) if rec_desc else "")

        w(16, row.get("H1", ""))
        w(17, row.get("Recommended H1", ""))

        fs = row.get("First Sentence after H1", "")
        w(18, fs)
        w(19, len(fs) if fs else "")

        rec_fs = row.get("Recommended First Sentence", "")
        w(20, rec_fs)
        w(21, len(rec_fs) if rec_fs else "")

        w(22, row.get("Notes", ""))

        ws.row_dimensions[ri].height = 60

    ws.freeze_panes = "A3"

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Score guide markdown ───────────────────────────────────────────────────────

SCORE_GUIDE = """
<div style="background:rgba(226,26,107,0.08);border-left:4px solid #E21A6B;
            padding:10px 14px;border-radius:4px;font-size:13px;margin:12px 0;">
  <strong>Salience Score Guide</strong> &nbsp;—&nbsp;
  <span style="color:#16a34a;font-weight:600">0.60+</span>&nbsp;Excellent: entity clearly dominates &nbsp;
  <span style="color:#ca8a04;font-weight:600">0.40–0.59</span>&nbsp;Good: strong signal &nbsp;
  <span style="color:#ea580c;font-weight:600">0.20–0.39</span>&nbsp;Moderate: competing entities dilute focus &nbsp;
  <span style="color:#dc2626;font-weight:600">&lt; 0.20</span>&nbsp;Weak: entity is not the clear subject
</div>
"""

# ══════════════════════════════════════════════════════════════════════════════
# TOP-LEVEL TABS
# ══════════════════════════════════════════════════════════════════════════════

# ── Sidebar: user info + logout ───────────────────────────────────────────────
with st.sidebar:
    if _current_user.get("picture"):
        st.image(_current_user["picture"], width=48)
    st.markdown(f"**{_current_user['name']}**")
    st.caption(_current_user["email"])
    st.caption(f"Role: {_current_user['role'].capitalize()}")
    if st.button("Sign out", key="signout_main"):
        auth.logout()
    st.divider()
    if auth.is_manager_or_above():
        st.markdown("**History**")
        _history = db.get_history(_current_user["email"], _current_user["role"])
        if _history:
            for _h in _history[:10]:
                _dl_name, _dl_data = db.get_export_file(_h["id"], _current_user["email"], _current_user["role"])
                if _dl_data:
                    st.download_button(
                        f"⬇ {_h['filename'][:28]}…" if len(_h["filename"]) > 30 else f"⬇ {_h['filename']}",
                        data=_dl_data,
                        file_name=_dl_name,
                        key=f"sidebar_dl_{_h['id']}",
                    )
                    st.caption(f"{_h['user_email']} · {_h['created_at'][:16]}")
        else:
            st.caption("No exports yet.")
    else:
        st.markdown("**My Exports**")
        _history = db.get_history(_current_user["email"], "user")
        if _history:
            for _h in _history[:10]:
                _dl_name, _dl_data = db.get_export_file(_h["id"], _current_user["email"], "user")
                if _dl_data:
                    st.download_button(
                        f"⬇ {_h['filename'][:28]}…" if len(_h["filename"]) > 30 else f"⬇ {_h['filename']}",
                        data=_dl_data,
                        file_name=_dl_name,
                        key=f"sidebar_dl_{_h['id']}",
                    )
                    st.caption(_h["created_at"][:16])
        else:
            st.caption("No exports yet.")

tab_single, tab_bulk, tab_recs = st.tabs(["Single Analysis", "Bulk Analysis", "On-Page Recommendations"])

# ══════════════════════════════════════════════════════════════════════════════
# SINGLE ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════

with tab_single:

    # ── Input tabs ────────────────────────────────────────────────────────────
    input_tab_paste, input_tab_url = st.tabs(["Paste text", "Load from URL"])

    with input_tab_url:
        st.text_input("Page URL", key="page_url")

        if st.button("Load page elements", key="load_page_elements"):
            if not st.session_state["page_url"].strip():
                st.error("Paste a URL first.")
            else:
                with st.spinner("Fetching page and scoring elements…"):
                    try:
                        data = fetch_page_elements(st.session_state["page_url"].strip())
                        st.session_state["url_meta_title"] = data.get("meta_title", "")
                        st.session_state["url_meta_description"] = data.get("meta_description", "")
                        st.session_state["url_h1"] = data.get("h1", "")
                        st.session_state["url_first_sentence"] = data.get("first_sentence_after_h1", "")
                        st.session_state["original_text"] = data.get("prefill_original", "")

                        # Auto-score each element
                        elem_map = {
                            "url_meta_title": st.session_state["url_meta_title"],
                            "url_meta_description": st.session_state["url_meta_description"],
                            "url_h1": st.session_state["url_h1"],
                            "url_first_sentence": st.session_state["url_first_sentence"],
                        }
                        scores = {}
                        for key, text_val in elem_map.items():
                            if text_val.strip():
                                try:
                                    ents = analyze_text_salience(text_val)
                                    te, ts = top_entity(ents)
                                    scores[key] = {"entity": te, "score": ts}
                                except Exception:
                                    scores[key] = {"entity": None, "score": None}
                        st.session_state["url_element_scores"] = scores
                        st.rerun()
                    except Exception as e:
                        st.error(f"Failed to fetch URL: {e}")

        # Display elements with inline scores + Use as original buttons
        elem_display = [
            ("Meta title", "url_meta_title", False),
            ("Meta description", "url_meta_description", True),
            ("H1", "url_h1", False),
            ("First sentence after H1", "url_first_sentence", True),
        ]
        elem_scores = st.session_state.get("url_element_scores", {})

        for label, key, multiline in elem_display:
            text_val = st.session_state.get(key, "")
            score_info = elem_scores.get(key, {})
            col_field, col_score, col_btn = st.columns([5, 3, 2])

            with col_field:
                if multiline:
                    st.text_area(label, value=text_val, height=80, disabled=True)
                else:
                    st.text_input(label, value=text_val, disabled=True)

            with col_score:
                if score_info.get("entity"):
                    te = score_info["entity"]
                    ts = score_info["score"]
                    color = score_color(ts)
                    st.markdown(
                        f'<div style="padding-top:{"26" if not multiline else "6"}px;font-size:12px;color:#9ca3af;">'
                        f'Top entity:<br>'
                        f'<span style="color:{color};font-weight:600">{te}</span> '
                        f'<span style="color:{color};font-weight:600">({ts:.2f})</span></div>',
                        unsafe_allow_html=True,
                    )

            with col_btn:
                if text_val.strip():
                    is_selected = st.session_state.get("selected_original_key") == key
                    st.markdown(f'<div style="padding-top:{"24" if not multiline else "4"}px;">', unsafe_allow_html=True)
                    st.button("Analyze this element", key=f"use_{key}",
                              on_click=use_as_original, args=(key,))
                    if is_selected:
                        st.markdown(
                            '<div style="font-size:11px;color:#16a34a;font-weight:700;'
                            'background:rgba(22,163,74,0.12);border:1px solid #16a34a;'
                            'border-radius:4px;padding:2px 7px;margin-top:4px;text-align:center;">'
                            '✓ Active original</div>',
                            unsafe_allow_html=True,
                        )
                    st.markdown('</div>', unsafe_allow_html=True)

    with input_tab_paste:
        st.text_area(
            "Paste the original content you want to analyze:",
            height=100,
            key="original_text",
        )
        orig_wc = wc(st.session_state.get("original_text", ""))
        if orig_wc:
            st.caption(f"{orig_wc} words")

    # ── Variation containers ───────────────────────────────────────────────────
    st.markdown("---")
    orig_wc_for_delta = wc(st.session_state.get("original_text", ""))

    with st.container(border=True):
        st.markdown("#### Variation 1")
        st.text_input("Target entity for Variation 1", key="target_entity_1", placeholder="Type any entity name…")
        _df_for_picker = st.session_state.get("display_df")
        if _df_for_picker is not None and not _df_for_picker.empty:
            st.selectbox(
                "Or pick from last analysis:",
                options=[""] + _df_for_picker["Entity"].tolist(),
                format_func=lambda x: "— pick an entity —" if x == "" else x,
                key="entity_picker_1",
                on_change=pick_entity, args=(1,),
                label_visibility="collapsed",
            )
        st.selectbox(
            "Element to optimize", options=_SOURCE_OPTIONS, key="variation_source_1",
            help="URL elements are available after loading a URL above.",
        )
        st.text_area(
            "Variation 1 text (paste manually or click Generate below):",
            height=100, key="variation_text_1",
        )
        v1_wc = wc(st.session_state.get("variation_text_1", ""))
        if v1_wc:
            delta = v1_wc - orig_wc_for_delta if orig_wc_for_delta else 0
            delta_str = f" ({'+' if delta >= 0 else ''}{delta} vs original)" if orig_wc_for_delta else ""
            st.caption(f"{v1_wc} words{delta_str}")
        copy_button(st.session_state.get("variation_text_1", ""), key="v1")

    with st.container(border=True):
        st.markdown("#### Variation 2")
        st.text_input("Target entity for Variation 2", key="target_entity_2", placeholder="Type any entity name…")
        if _df_for_picker is not None and not _df_for_picker.empty:
            st.selectbox(
                "Or pick from last analysis:",
                options=[""] + _df_for_picker["Entity"].tolist(),
                format_func=lambda x: "— pick an entity —" if x == "" else x,
                key="entity_picker_2",
                on_change=pick_entity, args=(2,),
                label_visibility="collapsed",
            )
        st.selectbox(
            "Element to optimize", options=_SOURCE_OPTIONS, key="variation_source_2",
            help="URL elements are available after loading a URL above.",
        )
        st.text_area(
            "Variation 2 text (paste manually or click Generate below):",
            height=100, key="variation_text_2",
        )
        v2_wc = wc(st.session_state.get("variation_text_2", ""))
        if v2_wc:
            delta = v2_wc - orig_wc_for_delta if orig_wc_for_delta else 0
            delta_str = f" ({'+' if delta >= 0 else ''}{delta} vs original)" if orig_wc_for_delta else ""
            st.caption(f"{v2_wc} words{delta_str}")
        copy_button(st.session_state.get("variation_text_2", ""), key="v2")

    st.markdown("---")

    # ── Target keyword ─────────────────────────────────────────────────────────
    st.text_input(
        "Target SEO keyword (optional — suggests which entity to optimize for)",
        key="target_keyword", placeholder="e.g. luxury spa Georgia",
    )
    if st.session_state.get("keyword_suggestion"):
        st.info(
            f"💡 Suggested entity for **\"{st.session_state['target_keyword']}\"**: "
            f"**{st.session_state['keyword_suggestion']}**"
        )

    # ── Controls row ───────────────────────────────────────────────────────────
    ctrl_left, ctrl_right = st.columns([5, 1])
    with ctrl_left:
        st.radio("Generate with Claude for:", ["Variation 1", "Variation 2"],
                 horizontal=True, key="assign_to")
    with ctrl_right:
        st.markdown('<div style="padding-top:28px;">', unsafe_allow_html=True)
        st.button("Clear all", on_click=clear_all)
        st.markdown('</div>', unsafe_allow_html=True)

    col_gen, col_analyze, col_suggest = st.columns([2, 2, 2])

    with col_gen:
        st.button(
            "Generate with Claude",
            on_click=on_generate_with_claude,
            disabled=not bool(ANTHROPIC_API_KEY),
        )

    if st.session_state.get("claude_error"):
        st.error(st.session_state["claude_error"])
    if not ANTHROPIC_API_KEY:
        st.warning("ANTHROPIC_API_KEY is not set. Add it to .env and restart.")

    # ── Analyze ────────────────────────────────────────────────────────────────
    original_text = st.session_state.get("original_text", "")
    variation_text_1 = st.session_state.get("variation_text_1", "")
    variation_text_2 = st.session_state.get("variation_text_2", "")

    with col_analyze:
        analyze_clicked = st.button("Analyze")

    with col_suggest:
        suggest_clicked = st.button(
            "Suggest entity",
            disabled=not bool(ANTHROPIC_API_KEY) or not st.session_state.get("target_keyword", "").strip(),
            help="Requires a target keyword and at least one Analyze run.",
        )

    if suggest_clicked:
        df = st.session_state.get("display_df")
        kw = st.session_state.get("target_keyword", "").strip()
        if df is None:
            st.warning("Run Analyze first.")
        elif not kw:
            st.warning("Enter a target keyword first.")
        else:
            with st.spinner("Asking Claude…"):
                try:
                    suggestion = suggest_entity_for_keyword(df["Entity"].tolist(), kw)
                    st.session_state["keyword_suggestion"] = suggestion
                    st.rerun()
                except Exception as e:
                    st.error(f"Suggestion failed: {e}")

    if analyze_clicked:
        v1_source_label = st.session_state.get("variation_source_1", "Pasted text")
        v2_source_label = st.session_state.get("variation_source_2", "Pasted text")

        if not any([original_text.strip(), variation_text_1.strip(), variation_text_2.strip()]):
            st.error("Add some text to analyze first — paste text in a variation box or click 'Analyze this element' to set an original.")
        else:
            with st.spinner("Scoring with Google NLP…"):
                all_entities = {}
                if original_text.strip():
                    all_entities["Original"] = analyze_text_salience(original_text)
                if variation_text_1.strip():
                    all_entities["Variation 1"] = analyze_text_salience(variation_text_1)
                    st.session_state["analyze_v1_label"] = v1_source_label
                else:
                    st.session_state["analyze_v1_label"] = ""
                if variation_text_2.strip():
                    all_entities["Variation 2"] = analyze_text_salience(variation_text_2)
                    st.session_state["analyze_v2_label"] = v2_source_label
                else:
                    st.session_state["analyze_v2_label"] = ""

            rows_list = []
            unique_entities = set(e for ents in all_entities.values() for e in ents)
            for entity in unique_entities:
                row = {"Entity": entity, "Type": None,
                       "Original": None, "Variation 1": None, "Variation 2": None}
                salience_scores = []
                for version, ents in all_entities.items():
                    if entity in ents:
                        s = float(ents[entity]["Salience"])
                        row["Type"] = ents[entity]["Type"]
                        row[version] = s
                        salience_scores.append(s)
                row["Average Salience"] = float(np.mean(salience_scores)) if salience_scores else np.nan
                rows_list.append(row)

            comparison_df = pd.DataFrame(rows_list)

            if comparison_df.empty:
                st.session_state["display_df"] = None
                st.warning(
                    "No entities were found. This usually means the text is too short or too generic. "
                    "Try pasting a longer passage with named people, places, or organizations."
                )
            else:
                comparison_df = comparison_df.sort_values(by="Average Salience", ascending=False)
                new_df = comparison_df.drop(columns=["Average Salience"])
                st.session_state["display_df"] = new_df

                # Save to session history (newest first, max 10)
                st.session_state["session_history"] = ([{
                    "time": datetime.now().strftime("%H:%M:%S"),
                    "url": st.session_state.get("page_url", "") or "pasted text",
                    "entity_1": st.session_state.get("target_entity_1", ""),
                    "entity_2": st.session_state.get("target_entity_2", ""),
                    "df": new_df.copy(),
                }] + st.session_state.get("session_history", []))[:10]

    # ── Results ────────────────────────────────────────────────────────────────
    if st.session_state.get("display_df") is not None:
        display_df = st.session_state["display_df"]

        # Winner highlight
        winner_lines = []
        for col_name, entity_key in [("Variation 1", "target_entity_1"), ("Variation 2", "target_entity_2")]:
            entity = st.session_state.get(entity_key, "").strip()
            if not entity:
                continue
            match = display_df[display_df["Entity"].str.lower() == entity.lower()]
            if match.empty:
                continue
            row = match.iloc[0]
            score_cols_present = [c for c in ["Original", "Variation 1", "Variation 2"]
                                   if c in row.index and pd.notna(row[c])]
            if not score_cols_present:
                continue
            best_col = max(score_cols_present, key=lambda c: row[c])
            best_score = row[best_col]
            orig_score = row.get("Original") if "Original" in row.index and pd.notna(row.get("Original")) else None
            improvement = (
                f" — up from <strong>{orig_score:.2f}</strong> original"
                if orig_score and best_col != "Original" else ""
            )
            color = score_color(best_score)
            winner_lines.append(
                f'<span style="color:{color};font-weight:600">▲ {entity}</span>: '
                f'best in <strong>{best_col}</strong> at '
                f'<span style="color:{color};font-weight:600">{best_score:.2f}</span>{improvement}'
            )

        if winner_lines:
            st.markdown(
                '<div style="background:rgba(255,255,255,0.04);border-radius:6px;'
                'padding:10px 14px;font-size:13px;margin-bottom:4px;">'
                + " &nbsp;|&nbsp; ".join(winner_lines) + "</div>",
                unsafe_allow_html=True,
            )

        st.markdown(SCORE_GUIDE, unsafe_allow_html=True)

        # Build display table with element labels in column headers
        v1_lbl = st.session_state.get("analyze_v1_label", "")
        v2_lbl = st.session_state.get("analyze_v2_label", "")
        table_df = display_df.copy()
        col_rename = {}
        if v1_lbl and "Variation 1" in table_df.columns:
            col_rename["Variation 1"] = f"Variation 1 — {v1_lbl}"
        if v2_lbl and "Variation 2" in table_df.columns:
            col_rename["Variation 2"] = f"Variation 2 — {v2_lbl}"
        if col_rename:
            table_df = table_df.rename(columns=col_rename)

        score_cols = [c for c in table_df.columns if c in
                      ["Original"] + list(col_rename.values()) +
                      (["Variation 1"] if "Variation 1" in table_df.columns else []) +
                      (["Variation 2"] if "Variation 2" in table_df.columns else [])]
        score_cols = [c for c in table_df.columns if c not in ("Entity", "Type")]
        styled = table_df.style.map(style_score, subset=score_cols)

        st.dataframe(
            styled,
            width="stretch",
            hide_index=True,
            on_select=assign_selected_entity,
            selection_mode="single-row",
            key="entity_table",
            column_config={c: st.column_config.NumberColumn(format="%.2f") for c in score_cols},
        )

        export_df = display_df.copy()
        page_url = st.session_state.get("page_url", "")
        export_df.insert(0, "URL", page_url if page_url.strip() else "—")
        _csv_bytes = export_df.to_csv(index=False).encode("utf-8")
        _csv_fname = f"salience_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        db.save_export(_current_user["email"], _csv_fname, "salience", _csv_bytes)
        st.download_button(
            "Export to CSV",
            data=_csv_bytes,
            file_name=_csv_fname,
            mime="text/csv",
        )

    # ── Session history ────────────────────────────────────────────────────────
    history = st.session_state.get("session_history", [])
    if len(history) > 1:
        with st.expander(f"Session history ({len(history)} analyses)"):
            for i, entry in enumerate(history):
                url_label = entry["url"] if entry["url"] != "pasted text" else "Pasted text"
                entities_label = " / ".join(
                    filter(None, [entry.get("entity_1"), entry.get("entity_2")])
                ) or "—"
                st.markdown(f"**{entry['time']}** — {url_label} — _{entities_label}_")
                h_df = entry["df"]
                h_score_cols = [c for c in ["Original", "Variation 1", "Variation 2"] if c in h_df.columns]
                h_styled = h_df.style.map(style_score, subset=h_score_cols)
                st.dataframe(
                    h_styled,
                    width="stretch",
                    hide_index=True,
                    key=f"hist_{i}",
                    column_config={
                        "Original": st.column_config.NumberColumn(format="%.2f"),
                        "Variation 1": st.column_config.NumberColumn(format="%.2f"),
                        "Variation 2": st.column_config.NumberColumn(format="%.2f"),
                    },
                )
                if i < len(history) - 1:
                    st.divider()

# ══════════════════════════════════════════════════════════════════════════════
# BULK ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════

with tab_bulk:
    st.markdown("### Bulk URL Analysis")
    st.markdown(
        "Paste a list of URLs (one per line). The tool fetches each page, scores the H1 and first "
        "sentence, and returns the top entity and salience score for each. Optionally enter a target "
        "entity to see its specific score across all pages."
    )

    st.text_area("URLs (one per line):", height=180, key="bulk_urls_input",
                 placeholder="https://example.com/page-1\nhttps://example.com/page-2")
    st.text_input("Target entity (optional — scored across all URLs):",
                  key="bulk_target_entity", placeholder="e.g. Napa Valley")

    if st.button("Run Bulk Analysis", key="run_bulk"):
        urls = [u.strip() for u in st.session_state["bulk_urls_input"].splitlines() if u.strip()]
        if not urls:
            st.error("Paste at least one URL.")
        else:
            results = []
            progress_bar = st.progress(0, text="Starting…")

            for i, url in enumerate(urls):
                progress_bar.progress(i / len(urls), text=f"Processing {i + 1}/{len(urls)}: {url[:70]}…")
                try:
                    data = fetch_page_elements(url)
                    h1 = data.get("h1", "")
                    fs = data.get("first_sentence_after_h1", "")
                    target = st.session_state.get("bulk_target_entity", "").strip()

                    h1_entity, h1_score = (None, None)
                    if h1.strip():
                        h1_ents = analyze_text_salience(h1)
                        h1_entity, h1_score = top_entity(h1_ents)
                        target_h1_score = h1_ents.get(target, {}).get("Salience") if target else None
                    else:
                        h1_ents = {}
                        target_h1_score = None

                    fs_entity, fs_score = (None, None)
                    if fs.strip():
                        fs_ents = analyze_text_salience(fs)
                        fs_entity, fs_score = top_entity(fs_ents)
                        target_fs_score = fs_ents.get(target, {}).get("Salience") if target else None
                    else:
                        target_fs_score = None

                    row = {
                        "URL": url,
                        "H1": h1,
                        "H1 Top Entity": h1_entity or "—",
                        "H1 Top Score": round(h1_score, 2) if h1_score is not None else None,
                        "First Sentence Top Entity": fs_entity or "—",
                        "First Sentence Top Score": round(fs_score, 2) if fs_score is not None else None,
                        "Error": "",
                    }
                    if target:
                        row["H1 Target Score"] = round(target_h1_score, 2) if target_h1_score is not None else None
                        row["FS Target Score"] = round(target_fs_score, 2) if target_fs_score is not None else None

                    results.append(row)

                except Exception as e:
                    results.append({
                        "URL": url, "H1": "", "H1 Top Entity": "—", "H1 Top Score": None,
                        "First Sentence Top Entity": "—", "First Sentence Top Score": None,
                        "Error": str(e),
                    })

            progress_bar.progress(1.0, text=f"Done! Processed {len(urls)} URLs.")
            st.session_state["bulk_results_df"] = pd.DataFrame(results)

    if st.session_state.get("bulk_results_df") is not None:
        bulk_df = st.session_state["bulk_results_df"]
        # Hide Error column when there are no errors
        bulk_display = bulk_df.copy()
        if "Error" in bulk_display.columns and bulk_display["Error"].fillna("").eq("").all():
            bulk_display = bulk_display.drop(columns=["Error"])
        bulk_score_cols = [c for c in bulk_display.columns if "Score" in c]
        bulk_styled = bulk_display.style.map(style_score, subset=bulk_score_cols)

        st.dataframe(
            bulk_styled,
            width="stretch",
            hide_index=True,
            column_config={c: st.column_config.NumberColumn(format="%.2f") for c in bulk_score_cols},
        )

        _bulk_bytes = bulk_display.to_csv(index=False).encode("utf-8")
        _bulk_fname = f"bulk_salience_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        db.save_export(_current_user["email"], _bulk_fname, "bulk", _bulk_bytes)
        st.download_button(
            "Export Bulk Results to CSV",
            data=_bulk_bytes,
            file_name=_bulk_fname,
            mime="text/csv",
        )

# ══════════════════════════════════════════════════════════════════════════════
# ON-PAGE RECOMMENDATIONS
# ══════════════════════════════════════════════════════════════════════════════

with tab_recs:
    st.markdown("### On-Page Recommendations")
    st.markdown(
        "Enter your target pages and keywords. The tool fetches each page's current "
        "elements, generates salience-optimized recommendations with Claude, and exports "
        "a formatted spreadsheet matching the Propellic deliverable template."
    )

    # Initialize session state for this tab
    _recs_default_rows = [
        {"Existing URL": "", "Primary Keyword": "", "KD": None,
         "Volume": None, "Recommended URL": "", "Secondary Keywords": ""}
    ] * 5
    if "recs_input_df" not in st.session_state:
        st.session_state["recs_input_df"] = pd.DataFrame(_recs_default_rows)
    if "recs_results" not in st.session_state:
        st.session_state["recs_results"] = []
    if "recs_client_name" not in st.session_state:
        st.session_state["recs_client_name"] = ""

    st.text_input(
        "Client / project name (shown in export header):",
        key="recs_client_name",
        placeholder="e.g. The Kartrite",
    )

    st.markdown("#### Step 1 — Enter pages and keywords")
    st.caption(
        "Existing URL and Primary Keyword are required per row. "
        "KD, Volume, Recommended URL, and Secondary Keywords are optional."
    )

    edited_input = st.data_editor(
        st.session_state["recs_input_df"],
        key="recs_data_editor",
        num_rows="dynamic",
        column_config={
            "Existing URL":      st.column_config.TextColumn("Existing URL",      width="large"),
            "Primary Keyword":   st.column_config.TextColumn("Primary Keyword",   width="medium"),
            "KD":                st.column_config.NumberColumn("KD",              min_value=0, max_value=100, step=1),
            "Volume":            st.column_config.NumberColumn("Volume",          min_value=0, step=1),
            "Recommended URL":   st.column_config.TextColumn("Recommended URL",   width="medium"),
            "Secondary Keywords":st.column_config.TextColumn("Secondary Keywords",width="medium"),
        },
        use_container_width=True,
        hide_index=True,
    )
    st.session_state["recs_input_df"] = edited_input

    st.markdown("#### Step 2 — Generate recommendations")

    if st.button("Run On-Page Recommendations", key="run_recs"):
        if not ANTHROPIC_API_KEY:
            st.error("ANTHROPIC_API_KEY is required. Add it to .env and restart.")
        else:
            # Filter to rows that have both URL and keyword
            try:
                run_df = edited_input[
                    edited_input["Existing URL"].astype(str).str.strip().astype(bool) &
                    edited_input["Primary Keyword"].astype(str).str.strip().astype(bool)
                ]
            except Exception:
                run_df = pd.DataFrame()

            if run_df.empty:
                st.error("Fill in at least one row with an Existing URL and Primary Keyword.")
            else:
                results = []
                n = len(run_df)
                progress_bar = st.progress(0, text="Starting…")

                for i, (_, row) in enumerate(run_df.iterrows()):
                    url         = str(row["Existing URL"]).strip()
                    primary_kw  = str(row["Primary Keyword"]).strip()
                    secondary   = str(row.get("Secondary Keywords") or "").strip()
                    rec_url     = str(row.get("Recommended URL") or "").strip()
                    kd_val      = row.get("KD")
                    vol_val     = row.get("Volume")

                    progress_bar.progress(
                        i / n,
                        text=f"Processing {i + 1}/{n}: {url[:60]}…"
                    )

                    result = {
                        "Existing URL":      url,
                        "Primary Keyword":   primary_kw,
                        "KD":                kd_val,
                        "Volume":            vol_val,
                        "Recommended URL":   rec_url,
                        "Secondary Keywords":secondary,
                    }

                    try:
                        page_data  = fetch_page_elements(url)
                        title      = page_data.get("meta_title", "")
                        desc       = page_data.get("meta_description", "")
                        h1         = page_data.get("h1", "")
                        first_sent = page_data.get("first_sentence_after_h1", "")

                        result["Page Title"]               = title
                        result["Meta Description"]          = desc
                        result["H1"]                        = h1
                        result["First Sentence after H1"]   = first_sent

                        rec_title  = generate_onpage_recommendation("Meta title",                title,      primary_kw, secondary)
                        rec_desc   = generate_onpage_recommendation("Meta description",           desc,       primary_kw, secondary)
                        rec_h1     = generate_onpage_recommendation("H1",                         h1,         primary_kw, secondary)
                        rec_first  = generate_onpage_recommendation("First sentence after H1",    first_sent, primary_kw, secondary)
                        notes      = generate_seo_notes(primary_kw, title, rec_title, desc, rec_desc, h1, rec_h1, first_sent, rec_first)

                        result["Recommended Title"]              = rec_title
                        result["Recommended Meta Description"]    = rec_desc
                        result["Recommended H1"]                  = rec_h1
                        result["Recommended First Sentence"]      = rec_first
                        result["Notes"]                           = notes
                        result["Error"]                           = ""

                    except Exception as exc:
                        for k in ["Page Title", "Meta Description", "H1", "First Sentence after H1",
                                  "Recommended Title", "Recommended Meta Description",
                                  "Recommended H1", "Recommended First Sentence", "Notes"]:
                            result[k] = ""
                        result["Error"] = str(exc)

                    results.append(result)

                progress_bar.progress(1.0, text=f"Done! Processed {n} pages.")
                st.session_state["recs_results"] = results

    # ── Results ────────────────────────────────────────────────────────────────
    if st.session_state.get("recs_results"):
        results = st.session_state["recs_results"]

        st.markdown("#### Step 3 — Review results")

        # Build display DataFrame (condensed view)
        disp_cols = [
            "Existing URL", "Primary Keyword",
            "Page Title", "Recommended Title",
            "H1", "Recommended H1",
            "First Sentence after H1", "Recommended First Sentence",
            "Notes",
        ]
        disp_data = [{k: r.get(k, "") for k in disp_cols} for r in results]
        disp_df = pd.DataFrame(disp_data)

        st.dataframe(
            disp_df,
            width="stretch",
            hide_index=True,
            column_config={
                "Existing URL":                st.column_config.TextColumn(width="medium"),
                "Primary Keyword":             st.column_config.TextColumn(width="small"),
                "Page Title":                  st.column_config.TextColumn(width="large"),
                "Recommended Title":           st.column_config.TextColumn(width="large"),
                "H1":                          st.column_config.TextColumn(width="medium"),
                "Recommended H1":              st.column_config.TextColumn(width="medium"),
                "First Sentence after H1":     st.column_config.TextColumn(width="large"),
                "Recommended First Sentence":  st.column_config.TextColumn(width="large"),
                "Notes":                       st.column_config.TextColumn(width="large"),
            },
        )

        # Show any errors
        errors = [r for r in results if r.get("Error")]
        if errors:
            with st.expander(f"⚠ {len(errors)} page(s) had errors"):
                for r in errors:
                    st.markdown(f"**{r['Existing URL']}**: {r['Error']}")

        # Export
        st.markdown("#### Step 4 — Export")
        client_nm   = st.session_state.get("recs_client_name", "")
        excel_bytes = build_onpage_excel(results, client_name=client_nm)
        fname       = f"onpage_recommendations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        db.save_export(_current_user["email"], fname, "onpage", excel_bytes)
        st.download_button(
            "📥 Export to Excel (.xlsx)",
            data=excel_bytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
